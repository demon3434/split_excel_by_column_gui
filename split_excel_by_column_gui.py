import os
import sys
import threading
import time
from datetime import datetime
from collections import defaultdict
from copy import copy
from pathlib import Path
from queue import Empty, Queue
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    BaseTk = TkinterDnD.Tk
    HAS_TKDND = True
except Exception:
    DND_FILES = None
    BaseTk = tk.Tk
    HAS_TKDND = False


APP_TITLE = "Excel按列拆分"
APP_ICON = Path(__file__).with_name("logo_icon_hd.ico")
LOGO_IMAGE = Path(__file__).with_name("logo2.png")
SUPPORTED_EXCEL_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def safe_name(value):
    text = str(value)
    for ch in '/\\:*?"<>|':
        text = text.replace(ch, "_")
    text = text.strip()
    return text or "空值"


class Splitter:
    def __init__(
        self,
        file_path,
        sheet_name,
        split_col_index,
        header_row,
        output_base_name,
        output_dir,
        logger,
        apply_first_row_format=True,
    ):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.split_col_index = split_col_index
        self.header_row = header_row
        self.output_base_name = output_base_name
        self.output_dir = Path(output_dir)
        self.log = logger
        self.apply_first_row_format = apply_first_row_format

    def run(self):
        task_start_wall = datetime.now()
        task_start = time.perf_counter()
        self.log(f"任务开始: {task_start_wall.strftime('%Y-%m-%d %H:%M:%S')}")
        self.log(f"读取文件: {self.file_path}")

        stage_start = time.perf_counter()
        self.log("阶段[读取与分析] 开始")
        groups, row_count, max_col = self._group_rows()
        self.log(f"阶段[读取与分析] 完成，耗时 {time.perf_counter() - stage_start:.2f} 秒")
        if row_count == 0:
            raise ValueError("所选工作表没有数据行")
        if not groups:
            raise ValueError("拆分列数据为空，无法拆分")

        stage_start = time.perf_counter()
        self.log("阶段[样式与格式处理] 开始")
        header_snapshot = self._snapshot_header(max_col)
        data_number_formats = []
        if self.apply_first_row_format:
            data_number_formats = self._snapshot_first_data_row_formats(max_col)
            self.log("数据格式策略: 已启用（应用首行数据格式）")
        else:
            self.log("数据格式策略: 已禁用（使用Excel默认格式）")
        self.log(f"阶段[样式与格式处理] 完成，耗时 {time.perf_counter() - stage_start:.2f} 秒")

        self.log(f"数据行数: {row_count}")
        self.log(f"分组数量: {len(groups)}")

        output_dir = self.output_dir
        output_dir.mkdir(parents=True, exist_ok=True)

        total = len(groups)
        stage_start = time.perf_counter()
        self.log("阶段[拆分输出] 开始")
        for i, (key, rows) in enumerate(groups.items(), start=1):
            self.log(f"({i}/{total}) {key} -> {len(rows)} 行")

            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            self._apply_header_snapshot(ws, header_snapshot, max_col)

            for row_idx, row_values in enumerate(rows, start=self.header_row + 1):
                for col_idx, cell_value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    if self.apply_first_row_format and col_idx <= len(data_number_formats):
                        cell.number_format = data_number_formats[col_idx - 1]

            file_name = f"【{safe_name(key)}】{self.output_base_name}.xlsx"
            wb.save(output_dir / file_name)

        self.log(f"阶段[拆分输出] 完成，耗时 {time.perf_counter() - stage_start:.2f} 秒")
        self.log(f"完成，输出目录: {output_dir}")
        self.log(f"任务总耗时: {time.perf_counter() - task_start:.2f} 秒")
        return output_dir

    def _group_rows(self):
        groups = defaultdict(list)
        row_count = 0

        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            if self.sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表不存在: {self.sheet_name}")

            ws = wb[self.sheet_name]
            max_col = ws.max_column or 1
            if self.header_row < 1 or self.header_row > (ws.max_row or 1):
                raise ValueError("表头行号超出范围")
            if self.split_col_index >= max_col:
                raise ValueError("所选拆分列超出范围")

            for row in ws.iter_rows(min_row=self.header_row + 1, values_only=True):
                values = list(row) if row else []
                if len(values) < max_col:
                    values.extend([None] * (max_col - len(values)))
                elif len(values) > max_col:
                    values = values[:max_col]

                if all(v in (None, "") for v in values):
                    continue
                row_count += 1

                key = values[self.split_col_index]
                if key in (None, ""):
                    continue
                groups[key].append(values)

            return groups, row_count, max_col
        finally:
            wb.close()

    def _snapshot_header(self, max_col):
        wb = load_workbook(self.file_path, read_only=False, data_only=False)
        try:
            ws = wb[self.sheet_name]
            header_cells = []
            row_heights = {}
            col_widths = {}

            for row_idx in range(1, self.header_row + 1):
                if ws.row_dimensions[row_idx].height is not None:
                    row_heights[row_idx] = ws.row_dimensions[row_idx].height

                row_cells = []
                for col_idx in range(1, max_col + 1):
                    src = ws.cell(row=row_idx, column=col_idx)
                    row_cells.append(
                        {
                            "value": src.value,
                            "font": copy(src.font),
                            "fill": copy(src.fill),
                            "border": copy(src.border),
                            "alignment": copy(src.alignment),
                            "number_format": src.number_format,
                            "protection": copy(src.protection),
                        }
                    )
                header_cells.append(row_cells)

            for col_idx in range(1, max_col + 1):
                letter = get_column_letter(col_idx)
                width = ws.column_dimensions[letter].width
                if width is not None:
                    col_widths[letter] = width

            merged_ranges = []
            for merged in ws.merged_cells.ranges:
                if merged.min_row >= 1 and merged.max_row <= self.header_row:
                    merged_ranges.append(str(merged))

            return {
                "header_cells": header_cells,
                "row_heights": row_heights,
                "col_widths": col_widths,
                "merged_ranges": merged_ranges,
            }
        finally:
            wb.close()

    def _snapshot_first_data_row_formats(self, max_col):
        wb = load_workbook(self.file_path, read_only=False, data_only=False)
        try:
            ws = wb[self.sheet_name]
            start_row = self.header_row + 1
            if start_row > (ws.max_row or 0):
                return ["General"] * max_col

            for row_idx in range(start_row, (ws.max_row or 0) + 1):
                row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
                if all(v in (None, "") for v in row_values):
                    continue
                return [
                    ws.cell(row=row_idx, column=col_idx).number_format or "General"
                    for col_idx in range(1, max_col + 1)
                ]
            return ["General"] * max_col
        finally:
            wb.close()

    def _apply_header_snapshot(self, ws, snapshot, max_col):
        for letter, width in snapshot["col_widths"].items():
            ws.column_dimensions[letter].width = width

        for row_idx, row in enumerate(snapshot["header_cells"], start=1):
            if row_idx in snapshot["row_heights"]:
                ws.row_dimensions[row_idx].height = snapshot["row_heights"][row_idx]
            for col_idx in range(1, max_col + 1):
                style_data = row[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx, value=style_data["value"])
                cell.font = copy(style_data["font"])
                cell.fill = copy(style_data["fill"])
                cell.border = copy(style_data["border"])
                cell.alignment = copy(style_data["alignment"])
                cell.number_format = style_data["number_format"]
                cell.protection = copy(style_data["protection"])

        for merged in snapshot["merged_ranges"]:
            ws.merge_cells(merged)


class App(BaseTk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("900x620")
        self.minsize(820, 560)
        if APP_ICON.exists():
            self.iconbitmap(str(APP_ICON))

        self.file_var = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.column_var = tk.StringVar()
        self.header_row_var = tk.StringVar(value="1")
        self.name_var = tk.StringVar()
        self.output_dir_var = tk.StringVar(value=str(self._default_output_dir()))
        self.apply_first_row_format_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="就绪")

        self.columns_map = {}
        self.last_output_dir = None
        self.log_queue = Queue()
        self.worker = None
        self.meta_worker = None

        self.logo_img = None
        self.logo_img_raw = None
        self.logo_label = None
        self.logo_frame = None
        self._logo_visible = False
        self._compact_layout = False

        self._setup_styles()
        self._build_ui()
        self._enable_file_drop()
        self.after(120, self._drain_log_queue)

    def _setup_styles(self):
        style = ttk.Style()
        style.configure("Compact.TButton", padding=(6, 2))
        style.configure("Compact.TLabelframe", padding=8)
        style.configure("Compact.TEntry", padding=(4, 2))
        style.configure("Compact.TCombobox", padding=(4, 2))

    def _build_ui(self):
        root = ttk.Frame(self, padding=12)
        root.pack(fill=tk.BOTH, expand=True)

        self.form = ttk.LabelFrame(root, text="拆分参数", style="Compact.TLabelframe")
        self.form.pack(fill=tk.X)

        ttk.Label(self.form, text="Excel文件").grid(row=0, column=0, sticky=tk.W, padx=(0, 6), pady=3)
        self.file_entry = ttk.Entry(self.form, textvariable=self.file_var, style="Compact.TEntry", width=56)
        self.file_entry.grid(row=0, column=1, columnspan=5, sticky=tk.EW, pady=3)

        self.pick_btn = ttk.Button(self.form, text="选择", width=8, style="Compact.TButton", command=self._pick_file)
        self.pick_btn.grid(row=0, column=6, padx=(6, 0), pady=3, sticky=tk.EW)

        ttk.Label(self.form, text="工作表").grid(row=1, column=0, sticky=tk.W, padx=(0, 6), pady=3)
        self.sheet_combo = ttk.Combobox(
            self.form, textvariable=self.sheet_var, width=12, state="readonly", style="Compact.TCombobox"
        )
        self.sheet_combo.grid(row=1, column=1, sticky=tk.W, padx=(0, 0), pady=3)
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_changed)

        ttk.Label(self.form, text="表头行号").grid(row=1, column=2, sticky=tk.W, padx=(12, 6), pady=3)
        self.header_row_entry = ttk.Entry(self.form, textvariable=self.header_row_var, width=8, style="Compact.TEntry")
        self.header_row_entry.grid(row=1, column=3, sticky=tk.W, padx=(0, 0), pady=3)
        self.header_row_entry.bind("<FocusOut>", self._on_header_row_changed)
        self.header_row_entry.bind("<Return>", self._on_header_row_changed)

        ttk.Label(self.form, text="拆分列").grid(row=1, column=4, sticky=tk.W, padx=(12, 6), pady=3)
        self.col_combo = ttk.Combobox(
            self.form, textvariable=self.column_var, width=24, state="readonly", style="Compact.TCombobox"
        )
        self.col_combo.grid(row=1, column=5, sticky=tk.W, padx=(0, 0), pady=3)

        self.refresh_btn = ttk.Button(
            self.form, text="刷新", width=8, style="Compact.TButton", command=self._refresh_columns_by_header_row
        )
        self.refresh_btn.grid(row=1, column=6, padx=(6, 0), pady=3, sticky=tk.EW)

        ttk.Label(self.form, text="拆分后文件名").grid(row=2, column=0, sticky=tk.W, padx=(0, 6), pady=3)
        self.name_entry = ttk.Entry(self.form, textvariable=self.name_var, style="Compact.TEntry", width=24)
        self.name_entry.grid(row=2, column=1, columnspan=2, sticky=tk.EW, pady=3)

        self.apply_first_row_format_chk = ttk.Checkbutton(
            self.form,
            text="应用首行数据的格式",
            variable=self.apply_first_row_format_var,
        )
        self.apply_first_row_format_chk.grid(row=2, column=3, columnspan=4, sticky=tk.W, padx=(8, 0), pady=3)

        ttk.Label(self.form, text="保存文件夹").grid(row=3, column=0, sticky=tk.W, padx=(0, 6), pady=3)
        self.output_dir_entry = ttk.Entry(self.form, textvariable=self.output_dir_var, style="Compact.TEntry", width=56)
        self.output_dir_entry.grid(row=3, column=1, columnspan=5, sticky=tk.EW, pady=3)
        self.pick_output_btn = ttk.Button(
            self.form, text="选择", width=8, style="Compact.TButton", command=self._pick_output_dir
        )
        self.pick_output_btn.grid(row=3, column=6, padx=(6, 0), pady=3, sticky=tk.EW)

        self.logo_frame = ttk.Frame(self.form, width=1, height=1)
        self.logo_frame.grid(row=0, column=8, rowspan=3, padx=(8, 0), pady=2, sticky=tk.NE)
        self.logo_frame.grid_propagate(False)

        logo_path = self._resolve_logo_path()
        if logo_path is not None:
            try:
                self.logo_img_raw = tk.PhotoImage(file=str(logo_path))
                self.logo_label = ttk.Label(self.logo_frame)
                self.logo_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
                self._logo_visible = True
            except Exception:
                self.logo_label = ttk.Label(self.logo_frame, text="logo图片", anchor=tk.CENTER)
                self.logo_label.pack(fill=tk.BOTH, expand=True)
        else:
            self.logo_label = ttk.Label(self.logo_frame, text="logo图片", anchor=tk.CENTER)
            self.logo_label.pack(fill=tk.BOTH, expand=True)

        self.form.columnconfigure(1, weight=1)
        self.form.columnconfigure(7, weight=1)  # spacer column to push logo to the far right
        self.form.columnconfigure(8, minsize=1)
        self.form.bind("<Configure>", self._on_form_resize)
        self.after_idle(self._sync_logo_square)

        btn_row = ttk.Frame(root)
        btn_row.pack(fill=tk.X, pady=(8, 6))
        self.run_btn = ttk.Button(btn_row, text="开始拆分", width=10, style="Compact.TButton", command=self._start_split)
        self.run_btn.pack(side=tk.LEFT)
        self.open_btn = ttk.Button(btn_row, text="打开输出目录", width=12, style="Compact.TButton", command=self._open_output_dir)
        self.open_btn.pack(side=tk.LEFT, padx=(8, 0))

        log_group = ttk.LabelFrame(root, text="运行日志", padding=8)
        log_group.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(log_group, height=18, wrap=tk.WORD)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll = ttk.Scrollbar(log_group, orient=tk.VERTICAL, command=self.log_text.yview)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.configure(yscrollcommand=scroll.set, state=tk.DISABLED)

        self.progress = ttk.Progressbar(root, mode="indeterminate")
        self.progress.pack(fill=tk.X, pady=(8, 0))
        ttk.Label(root, textvariable=self.status_var, anchor=tk.W).pack(fill=tk.X, pady=(6, 0))

    def _pick_file(self):
        path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm"), ("全部文件", "*.*")],
        )
        if not path:
            return
        self._try_load_excel_path(path, source="选择")

    def _try_load_excel_path(self, path, source="拖拽"):
        p = Path(path)
        if not p.exists() or not p.is_file():
            self._append_log(f"[{source}] 文件不存在或不是文件: {path}")
            return
        if p.suffix.lower() not in SUPPORTED_EXCEL_EXTS:
            self._append_log(f"[{source}] 非Excel文件，已忽略: {path}")
            return
        self.file_var.set(str(p))
        self.name_var.set(p.stem)
        if not self.output_dir_var.get().strip():
            self.output_dir_var.set(str(self._default_output_dir()))
        self._append_log(f"[{source}] 已加载文件: {p}")
        self._load_sheets(str(p))

    def _enable_file_drop(self):
        if os.name != "nt":
            return
        if not HAS_TKDND:
            self._append_log("拖拽支持不可用：未安装 tkinterdnd2")
            return
        try:
            self.drop_target_register(DND_FILES)
            self.dnd_bind("<<Drop>>", self._on_tkdnd_drop)
            self._append_log("拖拽支持已启用：可直接将Excel文件拖到窗口")
        except Exception as exc:
            self._append_log(f"拖拽支持初始化失败: {exc}")

    def _on_tkdnd_drop(self, event):
        try:
            paths = list(self.tk.splitlist(event.data))
        except Exception:
            paths = [event.data] if getattr(event, "data", None) else []
        self.after(0, lambda p=paths: self._on_files_dropped(p))
        return "break"

    def _on_files_dropped(self, paths):
        if not paths:
            return
        if self.worker and self.worker.is_alive():
            self._append_log("拖拽已忽略：拆分任务正在运行")
            return
        if self.meta_worker and self.meta_worker.is_alive():
            self._append_log("拖拽已忽略：正在加载工作表，请稍候")
            return
        self._try_load_excel_path(paths[0], source="拖拽")

    def _pick_output_dir(self):
        initial_dir = self.output_dir_var.get().strip() or str(self._default_output_dir())
        path = filedialog.askdirectory(title="选择保存文件夹", initialdir=initial_dir)
        if path:
            self.output_dir_var.set(path)

    def _load_sheets(self, file_path):
        if self.meta_worker and self.meta_worker.is_alive():
            messagebox.showwarning("提示", "后台任务正在运行，请稍候")
            return

        self._reset_metadata_ui()
        self._set_busy("正在加载工作表...")

        def worker_task():
            try:
                wb = load_workbook(file_path, read_only=True, data_only=True)
                try:
                    sheets = wb.sheetnames
                    if not sheets:
                        raise ValueError("未读取到任何工作表")
                finally:
                    wb.close()
                self.log_queue.put(("sheets_success", sheets))
            except Exception as exc:
                self.log_queue.put(("meta_error", str(exc)))
            finally:
                self.log_queue.put(("meta_done", ""))

        self.meta_worker = threading.Thread(target=worker_task, daemon=True)
        self.meta_worker.start()

    def _on_sheet_changed(self, _event):
        self._refresh_columns_by_header_row()

    def _on_header_row_changed(self, _event):
        self._refresh_columns_by_header_row()

    def _refresh_columns_by_header_row(self):
        if self.meta_worker and self.meta_worker.is_alive():
            return
        file_path = self.file_var.get().strip()
        sheet = self.sheet_var.get().strip()
        if not file_path or not sheet:
            return

        try:
            header_row = self._parse_header_row()
        except ValueError as exc:
            messagebox.showerror("错误", str(exc))
            return

        self._set_busy("正在读取列名...")

        def worker_task():
            try:
                options, mapping = self._read_columns(file_path, sheet, header_row)
                self.log_queue.put(("columns_success", {"sheet": sheet, "options": options, "mapping": mapping}))
            except Exception as exc:
                self.log_queue.put(("meta_error", str(exc)))
            finally:
                self.log_queue.put(("meta_done", ""))

        self.meta_worker = threading.Thread(target=worker_task, daemon=True)
        self.meta_worker.start()

    def _read_columns(self, file_path, sheet_name, header_row):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表不存在: {sheet_name}")
            ws = wb[sheet_name]
            if header_row < 1 or header_row > (ws.max_row or 1):
                raise ValueError("表头行号超出范围")

            row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), None)
            if row is None:
                raise ValueError("读取表头失败")

            options = []
            mapping = {}
            for idx, name in enumerate(row):
                col_letter = get_column_letter(idx + 1)
                title = name if name not in (None, "") else f"列{idx + 1}"
                display = f"{col_letter} | {title}"
                options.append(display)
                mapping[display] = idx

            if not options:
                raise ValueError("所选表头行为空，无法读取列名")

            return options, mapping
        finally:
            wb.close()

    def _parse_header_row(self):
        raw = self.header_row_var.get().strip()
        if not raw:
            raise ValueError("请填写表头行号")
        if not raw.isdigit():
            raise ValueError("表头行号必须是正整数")
        value = int(raw)
        if value < 1:
            raise ValueError("表头行号必须大于等于 1")
        return value

    def _start_split(self):
        if self.worker and self.worker.is_alive():
            messagebox.showwarning("提示", "已有拆分任务在运行，请稍候")
            return
        if self.meta_worker and self.meta_worker.is_alive():
            messagebox.showwarning("提示", "正在加载元数据，请稍候")
            return

        file_path = self.file_var.get().strip()
        sheet_name = self.sheet_var.get().strip()
        col_display = self.column_var.get().strip()
        split_col_index = self.columns_map.get(col_display)
        output_base_name = safe_name(self.name_var.get().strip())
        output_dir = self.output_dir_var.get().strip() or str(self._default_output_dir())

        if not file_path:
            messagebox.showerror("错误", "请先选择 Excel 文件")
            return
        if not sheet_name:
            messagebox.showerror("错误", "请选择工作表")
            return
        if split_col_index is None:
            messagebox.showerror("错误", "请选择拆分列")
            return
        if not output_base_name:
            messagebox.showerror("错误", "请填写拆分后文件名")
            return
        if not output_dir:
            messagebox.showerror("错误", "请设置保存文件夹")
            return

        try:
            header_row = self._parse_header_row()
        except ValueError as exc:
            messagebox.showerror("错误", str(exc))
            return

        self._append_log("-" * 60)
        self._append_log(f"开始时刻: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self._set_busy("正在拆分...")

        def worker_task():
            try:
                splitter = Splitter(
                    file_path=file_path,
                    sheet_name=sheet_name,
                    split_col_index=split_col_index,
                    header_row=header_row,
                    output_base_name=output_base_name,
                    output_dir=output_dir,
                    logger=lambda msg: self.log_queue.put(("log", msg)),
                    apply_first_row_format=self.apply_first_row_format_var.get(),
                )
                out_dir = splitter.run()
                self.log_queue.put(("split_success", str(out_dir)))
            except Exception as exc:
                self.log_queue.put(("split_error", str(exc)))
            finally:
                self.log_queue.put(("split_done", ""))

        self.worker = threading.Thread(target=worker_task, daemon=True)
        self.worker.start()

    def _drain_log_queue(self):
        while True:
            try:
                event, content = self.log_queue.get_nowait()
            except Empty:
                break

            if event == "log":
                self._append_log(content)
            elif event == "sheets_success":
                sheets = content
                self.sheet_combo["values"] = sheets
                self.sheet_var.set(sheets[0])
                self._append_log(f"已加载工作表，共 {len(sheets)} 个")
                self._refresh_columns_by_header_row()
            elif event == "columns_success":
                data = content
                sheet = data["sheet"]
                options = data["options"]
                self.columns_map = data["mapping"]
                self.col_combo["values"] = options
                self.column_var.set(options[0] if options else "")
                self._append_log(f"工作表 [{sheet}] 列名加载完成，共 {len(options)} 列")
            elif event == "meta_error":
                self._append_log(f"读取失败: {content}")
                messagebox.showerror("错误", content)
            elif event == "meta_done":
                self._set_idle("就绪")
                self._sync_logo_square()
            elif event == "split_success":
                self.last_output_dir = content
                self._append_log(f"完成: {content}")
                messagebox.showinfo("完成", f"拆分完成\n输出目录: {content}")
            elif event == "split_error":
                self._append_log(f"错误: {content}")
                messagebox.showerror("错误", content)
            elif event == "split_done":
                self._set_idle("就绪")

        self.after(120, self._drain_log_queue)

    def _append_log(self, text):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"[{ts}] {text}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def _open_output_dir(self):
        if self.last_output_dir and Path(self.last_output_dir).exists():
            os.startfile(self.last_output_dir)
            return

        fallback = Path(self.output_dir_var.get().strip() or self._default_output_dir())
        if fallback.exists():
            os.startfile(str(fallback))
            return
        messagebox.showwarning("提示", "输出目录还不存在，请先运行拆分")

    def _reset_metadata_ui(self):
        self.sheet_combo["values"] = []
        self.col_combo["values"] = []
        self.sheet_var.set("")
        self.column_var.set("")
        self.columns_map = {}
        self.last_output_dir = None

    def _set_busy(self, text):
        self.status_var.set(text)
        self.run_btn.configure(state=tk.DISABLED)
        self.pick_btn.configure(state=tk.DISABLED)
        self.open_btn.configure(state=tk.DISABLED)
        self.refresh_btn.configure(state=tk.DISABLED)
        self.sheet_combo.configure(state="disabled")
        self.col_combo.configure(state="disabled")
        self.header_row_entry.configure(state="disabled")
        self.output_dir_entry.configure(state="disabled")
        self.pick_output_btn.configure(state=tk.DISABLED)
        self.apply_first_row_format_chk.configure(state=tk.DISABLED)
        self.progress.start(10)

    def _set_idle(self, text):
        self.status_var.set(text)
        self.run_btn.configure(state=tk.NORMAL)
        self.pick_btn.configure(state=tk.NORMAL)
        self.open_btn.configure(state=tk.NORMAL)
        self.refresh_btn.configure(state=tk.NORMAL)
        self.sheet_combo.configure(state="readonly")
        self.col_combo.configure(state="readonly")
        self.header_row_entry.configure(state="normal")
        self.output_dir_entry.configure(state="normal")
        self.pick_output_btn.configure(state=tk.NORMAL)
        self.apply_first_row_format_chk.configure(state=tk.NORMAL)
        self.progress.stop()

    def _on_form_resize(self, event):
        if not self._logo_visible:
            return
        should_compact = event.width < 900
        if should_compact != self._compact_layout:
            self._compact_layout = should_compact
            self._apply_compact_layout(should_compact)
        self._sync_logo_square()

    def _sync_logo_square(self):
        if not self.logo_frame:
            return
        # Use only the left-side controls to determine the 3-row height.
        # This avoids a feedback loop where logo size affects row height and keeps growing.
        row0_h = max(self.file_entry.winfo_reqheight(), self.pick_btn.winfo_reqheight())
        row1_h = max(
            self.sheet_combo.winfo_reqheight(),
            self.header_row_entry.winfo_reqheight(),
            self.col_combo.winfo_reqheight(),
        )
        row2_h = max(self.name_entry.winfo_reqheight(), self.apply_first_row_format_chk.winfo_reqheight())
        total_h = row0_h + row1_h + row2_h + 12  # include row paddings
        side = max(72, int(total_h))
        self.logo_frame.configure(width=side, height=side)
        self.form.columnconfigure(8, minsize=side)

        if self._logo_visible and self.logo_img_raw is not None and self.logo_label is not None:
            max_dim = max(self.logo_img_raw.width(), self.logo_img_raw.height())
            scale = max(1, (max_dim + side - 1) // side)
            self.logo_img = self.logo_img_raw.subsample(scale, scale)
            self.logo_label.configure(image=self.logo_img, text="")

    def _apply_compact_layout(self, compact):
        if compact:
            self.sheet_combo.configure(width=9)
            self.col_combo.configure(width=18)
            self.header_row_entry.configure(width=6)
            self.file_entry.configure(width=44)
            self.name_entry.configure(width=44)
            self.pick_btn.configure(width=7)
            self.refresh_btn.configure(width=7)
        else:
            self.sheet_combo.configure(width=12)
            self.col_combo.configure(width=24)
            self.header_row_entry.configure(width=8)
            self.file_entry.configure(width=56)
            self.name_entry.configure(width=56)
            self.pick_btn.configure(width=8)
            self.refresh_btn.configure(width=8)

    def _resolve_logo_path(self):
        candidates = []
        if getattr(sys, "frozen", False):
            meipass = getattr(sys, "_MEIPASS", None)
            if meipass:
                candidates.append(Path(meipass) / "logo2.png")
            candidates.append(Path(sys.executable).resolve().parent / "logo2.png")
        candidates.append(Path.cwd() / "logo2.png")
        candidates.append(LOGO_IMAGE)
        for path in candidates:
            if path.exists():
                return path
        return None

    def _default_output_dir(self):
        if getattr(sys, "frozen", False):
            return Path(sys.executable).resolve().parent / "拆分后"
        return Path(__file__).resolve().parent / "拆分后"


if __name__ == "__main__":
    App().mainloop()




